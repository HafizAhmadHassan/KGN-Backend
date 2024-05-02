import multiprocessing
import subprocess
import threading
import sys

import os
import pandas as pd
import smtplib, subprocess,time
import sys

from openpyxl import load_workbook
from openpyxl import Workbook

from kgnWebApp.models import AlarmSetting 
from django.contrib.auth.models import User
"""

We need to check this
https://stackoverflow.com/questions/2046603/is-it-possible-to-run-function-in-a-subprocess-without-threading-or-writing-a-se

"""
import re

# Define the string containing the pattern



import re




def parse_logs(latest_logs):
    """
    this is parse alarms and return me back

    """
    attributes_list=[]
    time_list=[]
    id_list=[]
    pattern = r"(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}),\d{3}\s+-\s+INFO\s+-\s+UPDATE\s+\w+\s+SET\s+(.*?)\s+FROM\s+\w+\s+WHERE\s+(\w+)\s*=\s*(\d+)"
    
    for log in latest_logs:


        # Use the findall() function from the re module to extract the pattern from the string
        matches = re.findall(pattern, log)

        date_time = matches[0][0]

        # Split the extracted attributes by ','
        attributes = [attr.split('=')[0].strip() for attr in matches[0][1].split(',')]

        # Extract the ID
        id_value = matches[0][3]

        # Print the extracted attribute names, ID, date, and time
        time_list.append(date_time)
        #print("Date and Time:", date_time)
        attributes_list.append(attributes)
        #print("Attribute Names:", attributes)
        id_list.append(id_value)
        #print("ID:", id_value)

    return attributes_list, time_list , id_list




def my_send_email():

    # Define content and header 
    # ----- Sending Customise --- #

    # Check Read.md if u need help
    with open('example.log', 'r') as file:
        lines = file.readlines()

    latest_log = lines[len(lines)-3:]
    """
    1. Parse logs Get : alarm name and Time machine id (done)
    2. Check from Alarm setting table if 'role' and 'sentemail' is True
    3. Also extract other things from 'AlarmSetting' table
    4. Send an email
    """
    attrs, dts, ids = parse_logs(latest_log)
    # [Header : Alarm Name + Header]
    # [Content : Code and T]

    
    mail=smtplib.SMTP('email-smtp.eu-central-1.amazonaws.com', 587)
    mail.ehlo()
    mail.starttls()
    sender='noreply@codexrsu4.it'
    mail.login('AKIATCMPCKJUU5YZP3VR','BNmNLMXdCkelGRlP0sZ5rHBXmdnFegLdqhmLaqZF2kpz')

    all_users = User.objects.all()
    
    for user in all_users:
        recipient=user.email
        header='To:'+ recipient +'\n'+'From:' \
        +sender+'\n'+'subject:testmail\n'
        content="Hello World"
        content=header+content
        mail.sendmail(sender, recipient, content)
    
    mail.close()



def copy_sheet(source_file, source_sheet_name, target_file, target_sheet_name):
    # Load the source workbook and sheet
    source_wb = load_workbook(source_file)
    source_sheet = source_wb[source_sheet_name]

    # Load the target workbook and sheet
    target_wb = load_workbook(target_file)
    target_sheet = target_wb[target_sheet_name]

    # Clear existing data in the target sheet
    target_sheet.delete_rows(1, target_sheet.max_row)

    # Copy the data
    for row in source_sheet.iter_rows(values_only=True):
        target_sheet.append(row)

    # Save the target workbook
    target_wb.save(target_file)

# Example usage




def read_machine(args):
    file_backup = os.path.join("", "book3_2_backup.xlsm")
    df_backup = pd.read_excel(file_backup,sheet_name='Sheet'+ args[1])

    if df_backup.empty:
        print("Sheet is empty")
        """
        1. calling subprocess in which insert machine id and and sheet values in database
        2. args[0]== id of machine args[1]== id of sheet , '1' means insert is true
        
        """
        #file_original = os.path.join("", "book3_2.xlsm")
        #df_original = pd.read_excel(file_original,'Sheet'+ args[1])
        
        #df_original.to_excel("book3_2_back.xlsm")
        
        copy_sheet('book3_2.xlsm', 'Sheet'+args[1], 'book3_2_backup.xlsm', 'Sheet'+args[1])
        print('Written to Excel File successfully.')


        subprocess.run(['python3', 'csv_to_mysql2.py',args[0],'1','Sheet'+ args[1]])
        print( "Machine Configured : ", args[0] )
    else:
        try:
            file_original = os.path.join("", "book3_2.xlsm")
            df_original = pd.read_excel(file_original,'Sheet'+ args[1])
        except PermissionError :
            time.sleep(1)
            file_original = os.path.join("", "book3_2.xlsm")
            df_original = pd.read_excel(file_original,'Sheet'+ args[1])

        if df_original.equals(df_backup) == False:
            copy_sheet('book3_2.xlsm', 'Sheet'+args[1], 'book3_2_backup.xlsm', 'Sheet'+args[1])
            # update the cloud
            subprocess.run(['python3', 'csv_to_mysql2.py',args[0],'0','Sheet'+ args[1]])
            # Storing the logs in text file
            # update set alarm name and then pass list
            #my_send_email(list specify the alarms and timing )
            

   
    
    
    """
    
    
    count=0
    num_Sheet=1

    while(1):
        # This will be only run for one machine
        try:
            file_new = os.path.join("", "book3.xlsm")
            df2 = pd.read_excel(file_new,'Sheet'+ args[2])
        except PermissionError :
            time.sleep(1)
            file_new = os.path.join("", "book3.xlsm")
            df2 = pd.read_excel(file_new,'Sheet'+ args[2])

        data_frame_same = df1.equals(df2)
        if data_frame_same == 1:
            # wait some files and check the files again
            #print("Its Same seceond :",count)
            if count == 0:
                #os.system(f"python3 csv_to_mysql2.py {sys.argv[1]} 1")
                
                subprocess.run(['python3', 'csv_to_mysql2.py',args[1],'1','Sheet'+ args[2]]) # 1 means for insertion
            #exec(open("csv_to_mysql2.py").read())
            count=1
            time.sleep(1)
            
        else:
            print("Yes Change")

            # problem is double
            subprocess.run(['python3', 'csv_to_mysql2.py',args[1],'0','Sheet'+ args[2]]) # 1 means for insertion

            count=count+1
            my_send_email()
            print("Email sent")

            time.sleep(1)
            df1=df2
            time.sleep(1)

"""











###############################
x = 0

def myfunc():
  global x
  x +=1
def run_script(arguments):
    read_machine(arguments)
    #subprocess.run(["python3", "csv_change.py"] + arguments)

if __name__ == "__main__":
    # List of arguments for the script

    # Initialize the script arguments list
    script_arguments = []

    # Define the number of argument sets per pattern
    sets_per_pattern = 1

    # Define the number of patterns
    num_patterns =10

    # Loop through each pattern
    for pattern_index in range(num_patterns):
        # Calculate the starting index for this pattern
        start_index = pattern_index * sets_per_pattern * 10 + 1
        
        # Generate argument sets for this pattern
        for i in range(sets_per_pattern):
            # Generate argument sets from 1 to 9 for each pattern
            for j in range(1, 11):
                script_arguments.append([str(int(sys.argv[1]) + start_index + i * 10 + j), str(j)])

    # Now script_arguments contains the desired 200 argument sets
    print(script_arguments[4:6])

    script_arguments = script_arguments[3:6]
#running 3 machines
c=1
while(1):
    for index in range(0,3):
        print(script_arguments[index])
        run_script(script_arguments[index])
        time.sleep(1)
        print(time, " : ",c)
        c+=1


"""



    
# Create a process pool
    with multiprocessing.Pool() as pool:
        # Run the script with different arguments in parallel
        pool.map(run_script, script_arguments)
# Create threads for each set of arguments

"""
