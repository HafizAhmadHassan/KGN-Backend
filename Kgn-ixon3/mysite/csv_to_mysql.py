"""
https://stackoverflow.com/questions/59229903/inserting-csv-into-mysql-database-with-python-library-mysql-connector


Pip3 install mysqlclient


https://stackoverflow.com/questions/76585758/mysqlclient-cannot-install-via-pip-cannot-find-pkg-config-name-in-ubuntu


Brew Install 

sudo apt-get install python3-dev default-libmysqlclient-dev build-essential



To Solve check Experiment There is dataType Tuple Issue
https://www.w3schools.com/python/python_mysql_insert.asp



If you want to put one single value then it would take a string

tname is string type
cursor.execute('INSERT INTO kgnWebApp_plcio(name) VALUES("%s")', 
    ((tname)))

    
but if you want more than 1 value then It would be tuple

cursor.execute('INSERT INTO kgnWebApp_plcio(name,unit, value) VALUES("%s", "%s", "%s")', 
          (tuple(row)))


"""


#import pandas as pd
import filecmp
from sqlalchemy import create_engine, types

import pymysql
import csv
import numpy as np
import pandas as pd
# Lets transform
import openpyxl

# Load the Excel workbook
workbook = openpyxl.load_workbook('book2.xlsx')

# Select the active worksheet (if it's not the first worksheet)
worksheet = workbook.active

# Access individual cells and extract values
cell_value_A1 = worksheet['M44'].value
print("Value in cell M44:", cell_value_A1)

# Alternatively, you can also use cell coordinates as follows:
# cell_value_A1 = worksheet.cell(row=1, column=1).value

# Access multiple cells in a range and extract values
list1 = []

for row in worksheet.iter_rows(min_row=44, max_row=93, min_col=15, max_col=15):
    for cell in row:
        #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
        #    print()
        if (cell.value is None) == False :
            #print("Hi",(cell.value))
            print('Hi',cell.value)
            list1.append(str(cell.value))

print(len(list1))
list2 = []

for row in worksheet.iter_rows(min_row=44, max_row=93, min_col=17, max_col=17):
    for cell in row:
        #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
        #    print()
        if (cell.value is None) == False :
            #print("Hi",(cell.value))
            list2.append((int(cell.value)) )

# Close the workbook


list3=tuple(list1+list2)
#print(pd.DataFrame(np.reshape(list, ( -1, 2))))
print("This is lenght",len(list3))
print(list3)

workbook.close()




db = pymysql.connect(host='database-1.c3gsasq6u7jz.eu-south-1.rds.amazonaws.com',
                     user='admin',password="12345678",
                     database='KGN_DB',
                     port=3306)

cursor = db.cursor()
cursor.execute('SELECT * FROM kgnWebApp_plcio')
# this does not work result = cursor.execute('SHOW COLUMNS FROM kgnWebApp_plcio')


result = cursor.execute('SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = "kgnWebApp_plcio"')
print("this is result",result)

#print("This is",result)


 

# Define a SQL query with %s placeholders

# Define values to be inserted





sql = "INSERT INTO kgnWebApp_plcio (unit_TimeOut_RollUp_Op,unit_TimeOut_RollUp_Cl,unit_TimeOut_Rot_Fw,unit_TimeOut_Rot_Bw,unit_TimeOut_Press_Fw,unit_TimeOut_Press_Bw,unit_TimeOutEndLoad,unit_Min_Weight,unit_Max_Weight,unit_Max_Weight_Standard_Deviation,unit_TImeCleaner,unit_Pulse_Cleaner,unit_Time_Winter_Cycle,unit_Full_80_Type_1,unit_Full_100_Type_1,unit_TimeFull,unit_TimeSlowPress,unit_TimeBuzzerLoad_1,unit_TimeBuzzerLoad_2,unit_TImeRotBw_AfterF_PressFW,unit_Set_Couter_Rotation,unit_TimeOut_Communication,unit_Time_Manut_Prev_RollUp,unit_Time_Manut_Prev_Crate,unit_Time_Manut_Prev_Pmp_Igien,unit_Time_Manut_Prev_Hydr,unit_Time_Manut_Prev_ValveFwBw,unit_Time_Manut_Prev_ValveSlow,unit_TimeSwitchOff,unit_TimeSwitchOff_PC,unit_NumeroPressate,unit_TypeWaste,unit_FunzioneTurista,unit_CicloClean,unit_PressatePesoMinoreSoglia,unit_WhiteListOpen,unit_Try,unit_CamEnabled,unit_TempoConteggioEncoder,unit_TempoFiltroZeroEncoder,unit_Full_80_Type_2,unit_Full_100_Type_2,unit_Full_80_Type_3,unit_Full_100_Type_3,unit_Buzzer_OFF,unit_Ripristino,unit_Full_80_Type_4,unit_Full_100_Type_4,unit_Full_80_Type_5,unit_Full_100_Type_5,value_TimeOut_RollUp_Op,value_TimeOut_RollUp_Cl,value_TimeOut_Rot_Fw,value_TimeOut_Rot_Bw,value_TimeOut_Press_Fw,value_TimeOut_Press_Bw,value_TimeOutEndLoad,value_Min_Weight,value_Max_Weight,value_Max_Weight_Standard_Deviation,value_TImeCleaner,value_Pulse_Cleaner,value_Time_Winter_Cycle,value_Full_80_Type_1,value_Full_100_Type_1,value_TimeFull,value_TimeSlowPress,value_TimeBuzzerLoad_1,value_TimeBuzzerLoad_2,value_TImeRotBw_AfterF_PressFW,value_Set_Couter_Rotation,value_TimeOut_Communication,value_Time_Manut_Prev_RollUp,value_Time_Manut_Prev_Crate,value_Time_Manut_Prev_Pmp_Igien,value_Time_Manut_Prev_Hydr,value_Time_Manut_Prev_ValveFwBw,value_Time_Manut_Prev_ValveSlow,value_TimeSwitchOff,value_TimeSwitchOff_PC,value_NumeroPressate,value_TypeWaste,value_FunzioneTurista,value_CicloClean,value_PressatePesoMinoreSoglia,value_WhiteListOpen,value_Try,value_CamEnabled,value_TempoConteggioEncoder,value_TempoFiltroZeroEncoder,value_Full_80_Type_2,value_Full_100_Type_2,value_Full_80_Type_3,value_Full_100_Type_3,value_Buzzer_OFF,value_Ripristino,value_Full_80_Type_4,value_Full_100_Type_4,value_Full_80_Type_5,value_Full_100_Type_5) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s"

# Execute the query with values
print("this is tuple \n",list3)

#cursor.execute(sql, list3)



cursor.execute('INSERT INTO kgnWebApp_plcio ( unit_TimeOut_RollUp_Op, unit_TimeOut_RollUp_Cl, unit_TimeOut_Rot_Fw, unit_TimeOut_Rot_Bw, unit_TimeOut_Press_Fw,\
    unit_TimeOut_Press_Bw, unit_TimeOutEndLoad, unit_Min_Weight, unit_Max_Weight, unit_Max_Weight_Standard_Deviation,\
    unit_TImeCleaner, unit_Pulse_Cleaner, unit_Time_Winter_Cycle, unit_Full_80_Type_1, unit_Full_100_Type_1,\
    unit_TimeFull, unit_TimeSlowPress, unit_TimeBuzzerLoad_1, unit_TimeBuzzerLoad_2, unit_TImeRotBw_AfterF_PressFW,\
    unit_Set_Couter_Rotation, unit_TimeOut_Communication, unit_Time_Manut_Prev_RollUp, unit_Time_Manut_Prev_Crate,\
    unit_Time_Manut_Prev_Pmp_Igien, unit_Time_Manut_Prev_Hydr, unit_Time_Manut_Prev_ValveFwBw, unit_Time_Manut_Prev_ValveSlow,\
    unit_TimeSwitchOff, unit_TimeSwitchOff_PC, unit_NumeroPressate, unit_TypeWaste, unit_FunzioneTurista, unit_CicloClean,\
    unit_PressatePesoMinoreSoglia, unit_WhiteListOpen, unit_Try, unit_CamEnabled, unit_TempoConteggioEncoder,\
    unit_TempoFiltroZeroEncoder, unit_Full_80_Type_2, unit_Full_100_Type_2, unit_Full_80_Type_3, unit_Full_100_Type_3,\
    unit_Buzzer_OFF, unit_Ripristino, unit_Full_80_Type_4, unit_Full_100_Type_4, unit_Full_80_Type_5, unit_Full_100_Type_5,\
    value_TimeOut_RollUp_Op, value_TimeOut_RollUp_Cl, value_TimeOut_Rot_Fw, value_TimeOut_Rot_Bw, value_TimeOut_Press_Fw,\
    value_TimeOut_Press_Bw, value_TimeOutEndLoad, value_Min_Weight, value_Max_Weight, value_Max_Weight_Standard_Deviation,\
    value_TImeCleaner, value_Pulse_Cleaner, value_Time_Winter_Cycle, value_Full_80_Type_1, value_Full_100_Type_1, value_TimeFull,\
    value_TimeSlowPress, value_TimeBuzzerLoad_1, value_TimeBuzzerLoad_2, value_TImeRotBw_AfterF_PressFW, value_Set_Couter_Rotation,\
    value_TimeOut_Communication, value_Time_Manut_Prev_RollUp, value_Time_Manut_Prev_Crate, value_Time_Manut_Prev_Pmp_Igien,\
    value_Time_Manut_Prev_Hydr, value_Time_Manut_Prev_ValveFwBw, value_Time_Manut_Prev_ValveSlow, value_TimeSwitchOff,\
    value_TimeSwitchOff_PC, value_NumeroPressate, value_TypeWaste, value_FunzioneTurista, value_CicloClean,\
    value_PressatePesoMinoreSoglia, value_WhiteListOpen, value_Try, value_CamEnabled, value_TempoConteggioEncoder,\
    value_TempoFiltroZeroEncoder, value_Full_80_Type_2, value_Full_100_Type_2, value_Full_80_Type_3, value_Full_100_Type_3,\
    value_Buzzer_OFF, value_Ripristino, value_Full_80_Type_4, value_Full_100_Type_4, value_Full_80_Type_5, value_Full_100_Type_5) \
    VALUES ( "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
          "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s")',list3)

#close the connection to the database.
db.commit()
cursor.close()












# MySQL connection details

"""
    cursor.execute('INSERT INTO testcsv(names, \
          classes, mark )' \
          'VALUES("%s", "%s", "%s")', 
          row)
"""


"""



import csv
import MySQLdb

mydb = MySQLdb.connect(host='localhost',
    user='root',
    passwd='',
    db='mydb')
cursor = mydb.cursor()

csv_data = csv.reader(file('students.csv'))
for row in csv_data:

    cursor.execute('INSERT INTO testcsv(names, \
          classes, mark )' \
          'VALUES("%s", "%s", "%s")', 
          row)
#close the connection to the database.
mydb.commit()
cursor.close()
print("Done")

"""




"""




import pymysql
import csv
db = pymysql.connect(host='database-1.c3gsasq6u7jz.eu-south-1.rds.amazonaws.com',
                     user='admin',password="12345678",
                     database='KGN_DB',
                     port=3306)

cursor = db.cursor()

cursor.execute('SELECT name FROM kgnWebApp_plcio')
csv_data = np.reshape(list, ( -1, 3))
print("=======================")


count =1
for ind in csv_data[:,0]:
    csv_data[:,0][count-1]=str(count)+' : '+ind +'\n'
    count=count+1

tname = np.array2string(csv_data[:5,0])

print(type(tname))



cursor.execute('INSERT INTO kgnWebApp_plcio(name) VALUES("%s")', 
    ((tname)))
#close the connection to the database.
db.commit()
cursor.close()

print(type(csv_data[:,2]))




"""



"""
for row in csv_data:
    print("=======")
    print(type(row) )
    print((tuple(row)))
    cursor.execute('INSERT INTO kgnWebApp_plcio(name,unit, value) VALUES("%s", "%s", "%s")', 
          (tuple(row)))
#close the connection to the database.
db.commit()
cursor.close()

"""

#hi


"""


csv_data = csv.reader(open('data1.csv'))
next(csv_data)
for row in csv_data:
    cursor.execute('INSERT INTO kgnWebApp_plcio(name,unit,vlaues) VALUES(%s, %s)',row)

db.commit()
cursor.close()

"""



"""

MYSQL_USERNAME = 'admin'
MYSQL_PASSWORD = '12345678'
MYSQL_HOST = 'database-1.c3gsasq6u7jz.eu-south-1.rds.amazonaws.com'
MYSQL_PORT = '3306'
MYSQL_DB_NAME = 'KGN_DB'

# Establish connection to MySQL
engine = create_engine(f'mysql+pymysql://{MYSQL_USERNAME}:{MYSQL_PASSWORD}@{MYSQL_HOST}:{MYSQL_PORT}/{MYSQL_DB_NAME}')

# Path to your CSV file
csv_file_path = 'path_to_your_csv_file.csv'


  
# save the dataframe as a csv file 

# Read CSV file into pandas DataFrame
data = pd.DataFrame(np.reshape(list, ( -1, 3)))

data.to_csv("data1.csv")

data = pd.read_csv("data1.csv",sep=',',quotechar='\'',encoding='utf8')
print("--------------------------------========-------------------")
print(data)
# Define table name in your database
table_name = 'kgnWebApp_plcio'

# Create table from DataFrame
data.to_sql(table_name, engine, index=False,if_exists='append')

# Commit changes

engine.dispose()

"""












