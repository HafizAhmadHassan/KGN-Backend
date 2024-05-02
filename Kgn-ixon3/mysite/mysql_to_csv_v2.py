


import pymysql
import ast
import numpy as np
import pandas as pd
# Lets transform
import openpyxl
from django.urls import path,include
import sys
import json







# Retrieve arguments passed from the calling script
table_name= sys.argv[1]
print("HI",sys.argv[2])
list1=ast.literal_eval(sys.argv[2])



print(f"Arguments received: {table_name}, {list1}")

workbook = openpyxl.load_workbook('book3_2.xlsm')
worksheet = workbook.active





print("This is list",list1)
count=0

if table_name == "PLCData":
    print("We are at ", table_name)
    for row in worksheet.iter_rows(min_row=4, max_row=19, min_col=5, max_col=5):
        for cell in row:
            #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
            #    print()
            if (cell.value is None) == False :
                #print("Hi",(cell.value))
                cell.value=list1[count]
                count=count+1



    for row in worksheet.iter_rows(min_row=4, max_row=19, min_col=11, max_col=11):
        for cell in row:
            #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
            #    print()
            if (cell.value is None) == False :
                #print("Hi",(cell.value))
                cell.value=list1[count]
                count=count+1


    for row in worksheet.iter_rows(min_row=4, max_row=19, min_col=17, max_col=17):
        for cell in row:
            #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
            #    print()
            if (cell.value is None) == False :
                #print("Hi",(cell.value))
                cell.value=list1[count]
                count=count+1




    for row in worksheet.iter_rows(min_row=24, max_row=39, min_col=5, max_col=5):
        for cell in row:
            #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
            #    print()
            if (cell.value is None) == False :
                #print("Hi",(cell.value))
                cell.value=list1[count]
                count=count+1



    for row in worksheet.iter_rows(min_row=24, max_row=39, min_col=11, max_col=11):
        for cell in row:
            #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
            #    print()
            if (cell.value is None) == False :
                #print("Hi",(cell.value))
                cell.value=list1[count]
                count=count+1

    # Close the workbook




    for row in worksheet.iter_rows(min_row=24, max_row=39, min_col=17, max_col=17):
        for cell in row:
            #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
            #    print()
            if (cell.value is None) == False :
                #print("Hi",(cell.value))
                cell.value=list1[count]
                count=count+1


    # Access multiple cells in a range and extract values
 

    for row in worksheet.iter_rows(min_row=44, max_row=59, min_col=5, max_col=5):
        for cell in row:
            #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
            #    print()
            if (cell.value is None) == False :
                #print("Hi",(cell.value))
                cell.value=list1[count]
                count=count+1


    for row in worksheet.iter_rows(min_row=44, max_row=59, min_col=11, max_col=11):
        for cell in row:
            #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
            #    print()
            if (cell.value is None) == False :
                #print("Hi",(cell.value))
                cell.value=list1[count]
                count=count+1


elif table_name=="PLCIO :":
# Access multiple cells in a range and extract values
    
    print('We are at PLCIO')

    for row in worksheet.iter_rows(min_row=44, max_row=93, min_col=15, max_col=15):
        for cell in row:
            #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
            #    print()
            if (cell.value is None) == False :
                #print("Hi",(cell.value))
                cell.value=list1[count]
                count=count+1



    for row in worksheet.iter_rows(min_row=44, max_row=93, min_col=17, max_col=17):
        for cell in row:
            #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
            #    print()
            if (cell.value is None) == False :
                #print("Hi",(cell.value))
                cell.value=list1[count]
                count=count+1





# Close the workbook
workbook.save("book3_2.xlsm")
workbook.close()










# this does not work result = cursor.execute('SHOW COLUMNS FROM kgnWebApp_plcio')




#print("This is",result)


 

# Define a SQL query with %s placeholders

# Define values to be inserted






# Execute the query with values
#print("this is tuple \n",list3)

#cursor.execute(sql, list3)

"""



#cursor.execute('INSERT INTO kgnWebApp_plcio ( unit_TimeOut_RollUp_Op, unit_TimeOut_RollUp_Cl, unit_TimeOut_Rot_Fw, unit_TimeOut_Rot_Bw, unit_TimeOut_Press_Fw,\
    unit_TimeOut_Press_Bw, unit_TimeOutEndLoad, unit_Min_Weight, unit_Max_Weight, unit_Max_Weight_Standard_Deviation,\
    unit_TImeCleaner, unit_Pulse_Cleaner, unit_Time_Winter_Cycle, unit_Full_80_Type_1, unit_Full_100_Type_1,\
    unit_TimeFull, unit_TimeSlowPress, unit_TimeBuzzerLoad_1, unit_TimeBuzzerLoad_2, unit_TImeRotBw_AfterF_PressFW,\
    unit_Set_Couter_Rotation, unit_TimeOut_Communication, unit_Time_Manut_Prev_RollUp, unit_Time_Manut_Prev_Crate,\
    unit_Time_Manut_Prev_Pmp_Igien, unit_Time_Manut_Prev_Hydr, unit_Time_Manut_Prev_ValveFwBw, unit_Time_Manut_Prev_ValveSlow,\
    unit_TimeSwitchOff, unit_TimeSwitchOff_PC, unit_NumeroPressate, unit_TypeWaste, unit_FunzioneTurista, unit_CicloClean,\
    unit_PressatePesoMinoreSoglia, unit_WhiteListOpen, unit_Try, unit_CamEnabled, unit_TempoConteggioEncoder,\
    unit_TempoFiltroZeroEncoder, unit_Full_80_Type_2, unit_Full_100_Type_2, unit_Full_80_Type_3, unit_Full_100_Type_3,\
    unit_Buzzer_OFF, unit_Ripristino, unit_Full_80_Type_4, unit_Full_100_Type_4, unit_Full_80_Type_5, unit_Full_100_Type_5,\
    value_TimeOut_RollUp_Op, value_TimeOut_RollUp_Cl, value_TimeOut_Rot_Fw, value_TimeOut_Rot_Bw, value_TimeOut_Press_Fw,\
    value_TimeOut_Press_Bw, value_TimeOutEndLoad, value_Min_Weight, value_Max_Weight, value_Max_Weight_Standard_Deviation,\
    value_TImeCleaner, value_Pulse_Cleaner, value_Time_Winter_Cycle, value_Full_80_Type_1, value_Full_100_Type_1, value_TimeFull,\
    value_TimeSlowPress, value_TimeBuzzerLoad_1, value_TimeBuzzerLoad_2, value_TImeRotBw_AfterF_PressFW, value_Set_Couter_Rotation,\
    value_TimeOut_Communication, value_Time_Manut_Prev_RollUp, value_Time_Manut_Prev_Crate, value_Time_Manut_Prev_Pmp_Igien,\
    value_Time_Manut_Prev_Hydr, value_Time_Manut_Prev_ValveFwBw, value_Time_Manut_Prev_ValveSlow, value_TimeSwitchOff,\
    value_TimeSwitchOff_PC, value_NumeroPressate, value_TypeWaste, value_FunzioneTurista, value_CicloClean,\
    value_PressatePesoMinoreSoglia, value_WhiteListOpen, value_Try, value_CamEnabled, value_TempoConteggioEncoder,\
    value_TempoFiltroZeroEncoder, value_Full_80_Type_2, value_Full_100_Type_2, value_Full_80_Type_3, value_Full_100_Type_3,\
    value_Buzzer_OFF, value_Ripristino, value_Full_80_Type_4, value_Full_100_Type_4, value_Full_80_Type_5, value_Full_100_Type_5) \
    VALUES ( "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
          "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s",\
    "%s", "%s","%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s")',list3)

#close the connection to the database.
#db.commit()
#cursor.close()
"""











# MySQL connection details





#hi
















"""

with connection.cursor() as cursor:
    cursor.execute("select a, b, c from bar")
    print(cursor.fetchall())




import pandas as pd

# Create a Pandas Excel writer using XlsxWriter as the engine outside the loop.
writer = pd.ExcelWriter('pandas_simple.xlsx', engine='xlsxwriter')
# Sample loop, replace with directory browsing loop
for i in range(7):
    # Sample Pandas dataframe. Replace with SQL query and resulting data frame.
    df = pd.DataFrame({'DataFromSQLQuery': ['SQL query result {0}'.format(i)]})
    # Convert the dataframe to an XlsxWriter Excel object.
    df.to_excel(writer, sheet_name='Sheet{0}'.format(i))
# Close the Pandas Excel writer and output the Excel file.
writer.save()

"""