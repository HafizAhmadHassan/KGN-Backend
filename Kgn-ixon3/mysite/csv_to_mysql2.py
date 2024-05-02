
#import pandas as pd

import pymysql
# Lets transform
import openpyxl
import sys
import time

import logging

# Configure logging
logging.basicConfig(filename='example.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')



# Retrieve Django admin users

# Print the usernames of admin users



def convert_insert_to_update(insert_query,id):
    # Parse the INSERT query to extract table name, columns, and values
    #insert_query = insert_query.lower()  # Convert to lowercase for case insensitivity
    insert_query = insert_query.replace('INSERT INTO ', '')  # Remove 'INSERT INTO ' prefix
    table_name_end_index = insert_query.find('(')
    table_name = insert_query[:table_name_end_index].strip()
    columns_and_values = insert_query[table_name_end_index + 1: -1].strip()  # Strip parentheses

    columns, values = map(str.strip, columns_and_values.split('VALUES'))
    columns = columns.strip('(').strip(')').split(',')
    values = values.strip('(').strip(')').split(',')

    # Construct the UPDATE query
    update_query = f"UPDATE {table_name} SET "
    for column, value in zip(columns, values):
        update_query += f"{column.strip()} = {value.strip()}, "

    update_query = update_query.rstrip(', ')  # Remove the trailing comma
    update_query += ' WHERE id = ' + id
    update_query += ";"  # Add a semicolon at the end

    return update_query






id = sys.argv[1]
insert = sys.argv[2]
print("this is ID Save it ",id)
time.sleep(1)
# Load the Excel workbook
workbook = openpyxl.load_workbook('book3_2.xlsm')

# Select the active worksheet (if it's not the first worksheet)
#worksheet = workbook.active

worksheet = workbook[str(sys.argv[3])]

# Access individual cells and extract values
cell_value_A1 = worksheet['M44'].value
print("Value in cell M44:", cell_value_A1)

# Alternatively, you can also use cell coordinates as follows:
# cell_value_A1 = worksheet.cell(row=1, column=1).value

# Access multiple cells in a range and extract values
list1 = []
list1.append(int(id))

for row in worksheet.iter_rows(min_row=4, max_row=19, min_col=5, max_col=5):
    for cell in row:
        #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
        #    print()
        if (cell.value is None) == False :
            #print("Hi",(cell.value))
            list1.append(int(cell.value))

print(len(list1))
list2 = []

for row in worksheet.iter_rows(min_row=4, max_row=19, min_col=11, max_col=11):
    for cell in row:
        #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
        #    print()
        if (cell.value is None) == False :
            #print("Hi",(cell.value))
            list2.append((int(cell.value)) )

# Close the workbook


list3 = []

for row in worksheet.iter_rows(min_row=4, max_row=19, min_col=17, max_col=17):
    for cell in row:
        #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
        #    print()
        if (cell.value is None) == False :
            #print("Hi",(cell.value))
            list3.append((int(cell.value)) )



# Access multiple cells in a range and extract values
list4 = []

for row in worksheet.iter_rows(min_row=24, max_row=39, min_col=5, max_col=5):
    for cell in row:
        #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
        #    print()
        if (cell.value is None) == False :
            #print("Hi",(cell.value))

            list4.append(int(cell.value))

print(len(list4))
list5 = []

for row in worksheet.iter_rows(min_row=24, max_row=39, min_col=11, max_col=11):
    for cell in row:
        #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
        #    print()
        if (cell.value is None) == False :
            #print("Hi",(cell.value))
            list5.append((int(cell.value)) )

# Close the workbook


list6 = []

for row in worksheet.iter_rows(min_row=24, max_row=39, min_col=17, max_col=17):
    for cell in row:
        #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
        #    print()
        if (cell.value is None) == False :
            #print("Hi",(cell.value))
            list6.append((int(cell.value)) )


# Access multiple cells in a range and extract values
list7 = []

for row in worksheet.iter_rows(min_row=44, max_row=59, min_col=5, max_col=5):
    for cell in row:
        #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
        #    print()
        if (cell.value is None) == False :
            #print("Hi",(cell.value))
            list7.append(int(cell.value))

print(len(list7))

list8 = []

for row in worksheet.iter_rows(min_row=44, max_row=59, min_col=11, max_col=11):
    for cell in row:
        #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
        #    print()
        if (cell.value is None) == False :
            #print("Hi",(cell.value))
            list8.append((int(cell.value)) )


list9 = []
list9.append(id)
for row in worksheet.iter_rows(min_row=44, max_row=93, min_col=15, max_col=15):
    for cell in row:
        #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
        #    print()
        if (cell.value is None) == False :
            #print("Hi",(cell.value)
         
            list9.append(str(cell.value).replace('(','u').replace(')','u').replace('/','u'))


list10 = []

for row in worksheet.iter_rows(min_row=44, max_row=93, min_col=17, max_col=17):
    for cell in row:
        #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
        #    print()
        if (cell.value is None) == False :
            #print("Hi",(cell.value))
            list10.append(str(int(cell.value)) )





list11 = []
list11.append(id)
for row in worksheet.iter_rows(min_row=74, max_row=89, min_col=5, max_col=5):
    for cell in row:
        #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
        #    print()
        if (cell.value is None) == False :
            #print("Hi",(cell.value)
         
            list11.append(str(int(cell.value)))


list12 = []

for row in worksheet.iter_rows(min_row=91, max_row=99, min_col=5, max_col=5):
    for cell in row:
        #if(isinstance(cell.value, int) or isinstance(cell.value, str)):
        #    print()
        if (cell.value is None) == False :
            #print("Hi",(cell.value))
            list12.append(str(int(cell.value)) )



# Close the workbook
workbook.close()




db = pymysql.connect(host='database-1.c3gsasq6u7jz.eu-south-1.rds.amazonaws.com',
                     user='admin',password="12345678",
                     database='KGN_DB',
                     port=3306)

cursor = db.cursor()
# this does not work result = cursor.execute('SHOW COLUMNS FROM kgnWebApp_plcio')

# This to insert plciodata
cursor.execute('SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = "kgnWebApp_plcdata"')
result_set = cursor.fetchall()
t=""
s=""

count=0
for vals in result_set:
    t=t+str(vals[0])+','
    s=s+"%s,"
    count+=1

# additional for id
s=s+"%s,"
count+=1

if insert == '1':
    t=str("INSERT INTO kgnWebApp_plcdata ("+'id,'+t[3:-1]+") VALUES ("+s[3:-1]+")")
    time.sleep(1)
    print("my count = ",count,len(tuple(list1+list2+list3+list4+list5+list6+list7+list8)))
    cursor.execute(t, tuple(list1+list2+list3+list4+list5+list6+list7+list8) )
    print("Insertion PLCData Success")
    logging.info(str(t))
    
elif insert == '0':
# Example usage:
    t=""
    s=""
    for vals in result_set:
        t=t+str(vals[0])+','
        s=s+"%s,"
    
    already_query = "SELECT * FROM kgnWebApp_plcdata" + " WHERE id = " + id
    cursor.execute(already_query)
    result_set=cursor.fetchall()
    #print("This is already existing row",list(result_set[0])[1:])
    mystring_prev=",".join(str(element) for element in list(result_set[0])[1:])
    insert_query_prev = str("INSERT INTO kgnWebApp_plcdata ("+t[3:-1]+") VALUES ("+mystring_prev+")")
    update_query_prev = convert_insert_to_update(insert_query_prev,id)
    #print("Prev Converted UPDATE query:")
    #print(update_query_prev)




    mystring=",".join(str(element) for element in list1[1:]+list2+list3+list4+list5+list6+list7+list8)
    insert_query = str("INSERT INTO kgnWebApp_plcdata ("+t[3:-1]+") VALUES ("+mystring+")")
    update_query = convert_insert_to_update(insert_query,id)
    #print("Converted UPDATE query:")
    #print(update_query)

    # Split the update_query by ','
    split_query_prev = update_query_prev.split(',')
    split_query = update_query.split(',')
    """


    # Print each part of the split query
    for part in split_query:
        print( part.strip() )
    #previous
    for part_prev in split_query_prev:
        print(part_prev.strip() )
    """

    # create a new Query
    ind = 0
    for part in split_query:
        if part.strip() == split_query_prev[ind].strip():
           split_query[ind] =''
        ind = ind +1
    #print(split_query_prev)
    #print(split_query)
    
    if split_query[0] != '':
        update_query_final = split_query[0] + ','
    else:
        update_query_final = "UPDATE kgnWebApp_plcdata SET"

    for elements in split_query[1:]:
            if elements != '':
                update_query_final = update_query_final  + elements + ','
    update_query_final = update_query_final[:-1] + " WHERE id = " +id + ';'
    
    print("Final Query is here 1")
    print(update_query_final)
    if "SE " not in update_query_final:
        cursor.execute(update_query_final)
        logging.info(str(update_query_final))
    else:
        logging.info('0000-00-00 00:00:00,000 - INFO - UPDATE NOTABLE SET z=0 FROM table2 WHERE id = 0')
    print("Update PLCData Success")
    


# ---------------- PLC Data  Finished-------------- #   




# This to insert plcio
cursor.execute('SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = "kgnWebApp_plcio"')
result_set = cursor.fetchall()
t=""
s=""
for vals in result_set:
    t=t+str(vals[0])+','
    s=s+"%s,"
# additional for id
s=s+"%s,"
if insert == '1':
    t=str("INSERT INTO kgnWebApp_plcio ("+'id,'+t[3:-1]+") VALUES (" +s[3:-1]+")")
    for index in range(len(list9+list10)):
        print("values ",list(list9+list10)[index]," : ",result_set[index])
    cursor.execute(t, tuple(list9+list10))
    print("Insertion PLCIO Success")
    logging.info(str(t))

    
elif insert == '0':
# Example usage:
    t=""
    s=""
    for vals in result_set:
        t=t+str(vals[0])+','
        s=s+"%s,"
    update_query_final=""
    already_query=""
    already_query = "SELECT * FROM kgnWebApp_plcio" + " WHERE id = " + id
    cursor.execute(already_query)
    result_set=cursor.fetchall()
    #print("This is already existing row",list(result_set[0])[1:])
    mystring_prev=""
    mystring_prev=",".join(str(element) for element in list(result_set[0])[1:])
    insert_query_prev=""
    insert_query_prev = str("INSERT INTO kgnWebApp_plcio ("+t[3:-1]+") VALUES ("+mystring_prev+")")
    update_query_prev=""
    update_query_prev = convert_insert_to_update(insert_query_prev,id)
    
    print("Prev Converted UPDATE query: 1")
    print(update_query_prev)



    mystring=""
    mystring=",".join(str(element) for element in list9[1:]+list10)
    insert_query=""
    insert_query = str("INSERT INTO kgnWebApp_plcio ("+t[3:-1]+") VALUES ("+mystring+")")
    update_query=""
    update_query = convert_insert_to_update(insert_query,id)
    print("Converted UPDATE query: 2")
    print(update_query)

    # Split the update_query by ','
    split_query_prev=""
    split_query_prev = update_query_prev.split(',')
    split_query=""
    split_query = update_query.split(',')
    
    """
    # Print each part of the split query
    for part in split_query:
        print( part.strip() )
    #previous
    for part_prev in split_query_prev:
        print(part_prev.strip() )
    """

    # create a new Query
    ind = 0
    for part in split_query:
        if part.strip() == split_query_prev[ind].strip():
           split_query[ind] =''
        ind = ind +1
    #print(split_query_prev)
    #print(split_query)
    update_query_final=""
    if split_query[0] != '':
        update_query_final = split_query[0] + ','
    else:
        update_query_final = "UPDATE kgnWebApp_plcio SET"

    for elements in split_query[1:]:
        if elements != '':
            update_query_final = update_query_final  + elements + ','
    update_query_final = update_query_final[:-1] + " WHERE id = " +id + ';'
    
    print("Final Query is here 3")
    print(update_query_final)
    print("Final Query is here 4")
    print(update_query_final[:-13-len(id)])
    if "SE " not in update_query_final:
        cursor.execute(update_query_final)
        logging.info(str(update_query_final))
    else:
        logging.info('0000-00-00 00:00:00,000 - INFO - UPDATE NOTABLE SET z=0 FROM table2 WHERE id = 0')
    print("Update PLCio Success")



    # Log some messages
# --------- PLC IO ---------------#


# This to insert plcstatus
cursor.execute('SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = "kgnWebApp_plcstatus"')
result_set = cursor.fetchall()
t=""
s=""
for vals in result_set:
    t=t+str(vals[0])+','
    s=s+"%s,"
# additional for id
s=s+"%s,"

if insert == '1':
    t=str("INSERT INTO kgnWebApp_plcstatus ("+'id,'+t[3:-1]+") VALUES (" +s[3:-1]+")")
    for index in range(len(list11+list12)):
        print("values ",list(list11+list12)[index]," : ",result_set[index])
    cursor.execute(t, tuple(list11+list12))
    print("Insertion PLCStatus Success")
    logging.info(str(t))


    
elif insert == '0':
# Example usage:
    t=""
    s=""
    for vals in result_set:
        t=t+str(vals[0])+','
        s=s+"%s,"
    
    already_query = "SELECT * FROM kgnWebApp_plcstatus" + " WHERE id = " + id
    cursor.execute(already_query)
    result_set=cursor.fetchall()
    #print("This is already existing row",list(result_set[0])[1:])
    mystring_prev=",".join(str(element) for element in list(result_set[0])[1:])
    insert_query_prev = str("INSERT INTO kgnWebApp_plcstatus ("+t[3:-1]+") VALUES ("+mystring_prev+")")
    update_query_prev = convert_insert_to_update(insert_query_prev,id)
    #print("Prev Converted UPDATE query:")
    #print(update_query_prev)




    mystring=",".join(str(element) for element in list11[1:]+list12)
    insert_query = str("INSERT INTO kgnWebApp_plcstatus ("+t[3:-1]+") VALUES ("+mystring+")")
    update_query = convert_insert_to_update(insert_query,id)

    split_query_prev = update_query_prev.split(',')
    split_query = update_query.split(',')

    # create a new Query
    ind = 0
    for part in split_query:
        if part.strip() == split_query_prev[ind].strip():
           split_query[ind] =''
        ind = ind +1
    
    if split_query[0] != '':
        update_query_final = split_query[0] + ','
    else:
        update_query_final = "UPDATE kgnWebApp_plcstatus SET"

    for elements in split_query[1:]:
            if elements != '':
                update_query_final = update_query_final  + elements + ','
    update_query_final = update_query_final[:-1] + " WHERE id = " +id + ';'
    
    print("Final Query is here 1")
    print(update_query_final)
    if "SE " not in update_query_final:
        cursor.execute(update_query_final)
        logging.info(str(update_query_final))
    else:
        logging.info('0000-00-00 00:00:00,000 - INFO - UPDATE NOTABLE SET z=0 FROM table2 WHERE id = 0')
    print("Update PLCStatus Success")

   
# ------------ PLC Status ---------- #



   

db.commit()
cursor.close()
# Define a SQL query with %s placeholders

# Define values to be inserted






# MySQL connection details






#hi















